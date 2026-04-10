import os, re, sys, json, time
from pathlib import Path
from datetime import datetime

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

API_KEY      = os.environ.get("API_KEY_GEMINI_FUSE")
PASTA_LOTE   = r"C:\Users\ctucunduva\fontes fuse\images\Artworks Beta 1"
PASTA_SCRIPT = r"C:\Users\ctucunduva\fontes fuse\Claude Cowork - Image Search"
MODELO_FLASH = "gemini-3-flash-preview"
EXTENSOES    = [".jpg", ".jpeg", ".png", ".webp", ".gif", ".JPG", ".PNG", ".JPEG"]
PROMPT = (
    "Please find interpretations of this artwork on the internet: {identificacao}. "
    "If you can't find the references for the work online, don't try to guess or make "
    "your own interpretation; simply report that you couldn't find them. "
    "Formatting: Do not use bold or italics (no ** or * or _ symbols). "
    "Numbered lists, indentation, and paragraph breaks are allowed."
)
NAO_ENCONTROU = [
    "unable to find", "could not find", "couldn't find", "not find", "not found",
    "no references", "no online references", "no interpretations", "cannot find",
    "can't find", "i was unable", "i could not", "i couldn't", "did not find",
    "don't have", "do not have", "no specific information", "no details",
    "no information", "no results", "no specific",
]

def extrair_id(nome):
    stem = Path(nome).stem
    stem = re.sub(r'^A_\s*', '', stem)
    stem = re.sub(r'^A\s+', '', stem)
    stem = re.sub(r'[_\s]+text\s*$', '', stem, flags=re.IGNORECASE)
    stem = re.sub(r'[_\s]+[\(\[]?\d+[\)\]]?\s*$', '', stem)
    return stem.strip()

# --- obras do lote ---
arquivos = sorted([f for f in Path(PASTA_LOTE).iterdir() if f.is_file() and f.suffix in EXTENSOES])
grupos = {}       # chave -> arquivo representativo
duplicatas = {}   # chave -> lista de arquivos agrupados (incluindo o representativo)
for f in arquivos:
    if 'text' in f.stem.lower() and f.stem.lower().endswith('text'):
        continue
    chave = extrair_id(f.name)
    if chave not in grupos:
        grupos[chave] = f
        duplicatas[chave] = [f.name]
    else:
        duplicatas[chave].append(f.name)

# --- historico: carrega todos os JSONs anteriores (analise_obras_*) ---
historico = {}
jsons = sorted(Path(PASTA_SCRIPT).glob("analise_obras_*.json"))
jsons = [j for j in jsons if "_parcial" not in j.name]
for jpath in jsons:
    with open(jpath, encoding="utf-8") as f:
        data = json.load(f)
    for r in data:
        if r.get("status") == "sucesso" and r.get("analise"):
            chave = r["identificacao"]
            if chave not in historico:
                historico[chave] = {**r, "_fonte_json": jpath.name}

# --- checkpoint do lote atual (retoma de onde parou) ---
checkpoint_lote = {}
checkpoints = sorted(Path(PASTA_SCRIPT).glob("beta1_*_parcial.json"), reverse=True)
if checkpoints:
    with open(checkpoints[0], encoding="utf-8") as f:
        cp_data = json.load(f)
    for r in cp_data:
        if r.get("status") == "sucesso":
            checkpoint_lote[r["identificacao"]] = r
    print(f"Retomando checkpoint: {checkpoints[0].name} ({len(checkpoint_lote)} obras ja processadas)\n")

reuso = {k: historico[k] for k in grupos if k in historico and k not in checkpoint_lote}
novos = [(k, grupos[k]) for k in grupos if k not in historico and k not in checkpoint_lote]

print(f"Beta 1: {len(grupos)} obras unicas")
print(f"  Reuso do historico: {len(reuso)}")
print(f"  Novas (Gemini):     {len(novos)}")
print(f"  Modelo:             {MODELO_FLASH}\n")

# --- configura gemini ---
from google import genai
from google.genai import types
client = genai.Client(api_key=API_KEY)

def chamar(identificacao, caminho, modelo, timeout=120):
    with open(caminho, "rb") as f2:
        dados = f2.read()
    ext = Path(caminho).suffix.lower()
    mime = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
            ".webp": "image/webp", ".gif": "image/gif"}.get(ext, "image/jpeg")
    resp = client.models.generate_content(
        model=modelo,
        contents=[types.Part.from_bytes(data=dados, mime_type=mime),
                  PROMPT.format(identificacao=identificacao)],
        config=types.GenerateContentConfig(
            tools=[types.Tool(google_search=types.GoogleSearch())],
            http_options=types.HttpOptions(timeout=timeout * 1000)
        )
    )
    return resp.text

# --- monta resultados ---
# Usa timestamp do checkpoint se existir, para sobrescrever o mesmo arquivo
if checkpoints:
    timestamp = checkpoints[0].stem.replace("beta1_", "").replace("_parcial", "")
else:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
pasta_script = Path(PASTA_SCRIPT)
resultados   = list(checkpoint_lote.values())  # começa com o que já foi processado
idx          = len(resultados)
# 1) reuso
for chave, r in reuso.items():
    idx += 1
    resultados.append({
        "id": idx,
        "identificacao": chave,
        "arquivo": grupos[chave].name,
        "analise": r.get("analise", ""),
        "modelo_usado": r.get("modelo_usado", ""),
        "status": "sucesso",
        "timestamp": r.get("timestamp", ""),
        "fonte_dado": "historico",
        "fonte_data": r.get("timestamp", "")[:10],
        "fonte_json": r.get("_fonte_json", ""),
        "arquivos_agrupados": " | ".join(duplicatas.get(chave, [])),
    })
    print(f"[{idx}/{len(grupos)}] {chave[:65]}... REUSO ({r.get('_fonte_json', '')})")

# 2) novas
for chave, caminho in novos:
    idx += 1
    print(f"[{idx}/{len(grupos)}] {chave[:65]}...", end=" ", flush=True)
    try:
        texto = chamar(chave, str(caminho), MODELO_FLASH)
        print("OK")
        resultados.append({
            "id": idx, "identificacao": chave, "arquivo": caminho.name,
            "analise": texto, "modelo_usado": MODELO_FLASH, "status": "sucesso",
            "timestamp": datetime.now().isoformat(),
            "fonte_dado": "novo", "fonte_data": "", "fonte_json": "",
            "arquivos_agrupados": " | ".join(duplicatas.get(chave, [])),
        })
    except Exception as e:
        print(f"ERRO: {str(e)[:120]}")
        resultados.append({
            "id": idx, "identificacao": chave, "arquivo": caminho.name,
            "analise": "", "modelo_usado": "", "status": "erro",
            "erro": str(e), "timestamp": datetime.now().isoformat(),
            "fonte_dado": "novo", "fonte_data": "", "fonte_json": "",
            "arquivos_agrupados": " | ".join(duplicatas.get(chave, [])),
        })

    if idx % 5 == 0:
        cp = pasta_script / f"beta1_{timestamp}_parcial.json"
        with open(cp, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)
        print(f"  >> Checkpoint ({idx} obras)")
    time.sleep(3)

# --- salva JSON ---
print("\n" + "=" * 60)
saida_json = pasta_script / f"beta1_{timestamp}.json"
with open(saida_json, "w", encoding="utf-8") as f:
    json.dump(resultados, f, ensure_ascii=False, indent=2)
print(f"JSON: {saida_json}")

# --- salva Excel ---
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Beta 1"
cab = ["#", "Obra", "Arquivo", "Analise", "Modelo", "Status", "Timestamp", "Fonte", "Data Fonte", "JSON Origem", "Arquivos Agrupados (nao analisados)"]
hf    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hfont = Font(bold=True, color="FFFFFF", size=11)
for col, t in enumerate(cab, 1):
    c = ws.cell(row=1, column=col, value=t)
    c.fill = hf; c.font = hfont
    c.alignment = Alignment(horizontal="center", vertical="center")

alt_fill    = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
reuso_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
agrup_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # verde claro

# Expande resultados: cada arquivo agrupado vira linha propria
linhas_xlsx = []
for r in resultados:
    linhas_xlsx.append(r)
    ag = r.get("arquivos_agrupados", "")
    if ag:
        arquivos_ag = [x.strip() for x in ag.split("|") if x.strip()]
        arquivo_principal = r.get("arquivo", "")
        for arq in arquivos_ag:
            if arq == arquivo_principal:
                continue  # pula o proprio representativo (ja esta na linha principal)
            linhas_xlsx.append({
                "_tipo": "agrupado",
                "identificacao": r.get("identificacao", ""),
                "arquivo": arq,
                "analise": r.get("analise", ""),
                "modelo_usado": r.get("modelo_usado", ""),
                "status": "agrupado",
                "timestamp": r.get("timestamp", ""),
                "fonte_dado": f"agrupado com: {arquivo_principal}",
                "fonte_data": r.get("fonte_data", ""),
                "fonte_json": r.get("fonte_json", ""),
                "arquivos_agrupados": "",
            })

for i, r in enumerate(linhas_xlsx, 1):
    row  = i + 1
    tipo = r.get("_tipo", "")
    if tipo == "agrupado":
        fill = agrup_fill
    elif r.get("fonte_dado") == "historico":
        fill = reuso_fill
    else:
        fill = alt_fill if i % 2 == 0 else PatternFill()
    vals = [i, r.get("identificacao",""), r.get("arquivo",""), r.get("analise",""),
            r.get("modelo_usado",""), r.get("status",""), r.get("timestamp",""),
            r.get("fonte_dado",""), r.get("fonte_data",""), r.get("fonte_json",""),
            r.get("arquivos_agrupados","")]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top")

for col, w in enumerate([5, 45, 30, 100, 25, 10, 20, 25, 12, 30, 10], 1):
    ws.column_dimensions[get_column_letter(col)].width = w
for row in range(2, len(linhas_xlsx) + 2):
    ws.row_dimensions[row].height = 120
ws.freeze_panes = "A2"

saida_xlsx = pasta_script / f"beta1_{timestamp}.xlsx"
wb.save(saida_xlsx)
print(f"Excel: {saida_xlsx}")

suc  = sum(1 for r in resultados if r["status"] == "sucesso")
hist = sum(1 for r in resultados if r.get("fonte_dado") == "historico")
print(f"\nRESUMO: {suc}/{len(resultados)} OK | {hist} reuso | modelo: {MODELO_FLASH}")
