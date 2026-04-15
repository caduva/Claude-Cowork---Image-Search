"""
Processa obras do Banksy (linhas 488-511 do beta1.xlsx).
Envia imagem + prompt especifico para identificar cada obra individualmente.
Atualiza beta1.json e beta1.xlsx na pasta Revisado_14_04_2026.
"""
import os, re, sys, json, time
from pathlib import Path
from datetime import datetime

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import Alignment

API_KEY      = os.environ.get("API_KEY_GEMINI_FUSE")
PASTA_OUTPUT = Path(r"C:\Users\ctucunduva\fontes fuse\Claude Cowork - Image Search\Output Gemini Description\Revisado_14_04_2026")
PASTA_BANKSY = Path(r"C:\Users\ctucunduva\fontes fuse\Claude Cowork - Image Search\banksy")
MODELO       = "gemini-3-flash-preview"

# Prompt especifico: pede identificacao da obra pela imagem e busca referencias
PROMPT = """This is a work by the street artist Banksy.
Please identify the specific title or common name of this particular artwork based on the image provided.
Then find interpretations and references for this specific work on the internet.
Do not group it with other Banksy works — focus only on this individual piece.
If you cannot identify the specific work, describe what is depicted and search for the most similar known Banksy work.
If you truly cannot find any references for this specific image, report that clearly.
Formatting: Do not use bold or italics (no ** or * or _ symbols). Numbered lists, indentation, and paragraph breaks are allowed."""

FRASES_NAO = ['unable to find','could not find',"couldn't find",'i could not find',
    'no specific','no references','no information','no results','unable to locate',
    'cannot identify','could not identify']

def suc_rev(texto):
    t = (texto or '').lower()
    if not t: return 'nao'
    for f in FRASES_NAO:
        if f in t[:400]:
            return 'nao' if len(t) < 500 else 'parcial'
    return 'sim' if len(t) > 300 else 'parcial'

# lista de arquivos a processar (linhas 488-511, excluindo 487 que ja foi feita e 512)
ARQUIVOS_PLANILHA = [
    'A_Banksy_(16).jpg',
    'A_Banksy_1.jpg',
    'A_Banksy_12.jpeg',
    'A_Banksy_12.jpg',
    'A_Banksy_13.jpg',
    'A_Banksy_15.jpg',
    'A_Banksy_17.jpg',
    'A_Banksy_18.jpg',
    'A_Banksy_19.jpg',
    'A_Banksy_23.jpg',
    'A_Banksy_24.jpg',
    'A_Banksy_26.jpg',
    'A_Banksy_28.jpg',
    'A_Banksy_29.jpg',
    'A_Banksy_3.jpg',
    'A_Banksy_30.jpg',
    'A_Banksy_31.jpg',
    'A_Banksy_32.jpg',
    'A_Banksy_34.jpg',
    'A_Banksy_35.jpg',
    'A_Banksy_36.jpg',
    'A_Banksy_4.jpg',
    'A_Banksy_5.jpg',
    'A_Banksy_8.jpg',
]

client = genai.Client(api_key=API_KEY)

# carrega JSON
with open(PASTA_OUTPUT / 'beta1.json', encoding='utf-8') as f:
    principal = json.load(f)
with open(PASTA_OUTPUT / 'beta1_Obras_Sem_Referencia.json', encoding='utf-8') as f:
    sem_ref = json.load(f)

# indexa por arquivo
arquivo_para_reg_principal = {r.get('arquivo',''): r for r in principal}
arquivo_para_reg_semref    = {r.get('arquivo',''): r for r in sem_ref}

resultados = []

for i, nome_arquivo in enumerate(ARQUIVOS_PLANILHA, 1):
    img_path = PASTA_BANKSY / nome_arquivo
    if not img_path.exists():
        print(f'[{i}/{len(ARQUIVOS_PLANILHA)}] {nome_arquivo}: IMAGEM NAO ENCONTRADA')
        continue

    print(f'[{i}/{len(ARQUIVOS_PLANILHA)}] {nome_arquivo}...', end=' ', flush=True)

    try:
        with open(img_path, 'rb') as f2:
            dados = f2.read()
        ext  = img_path.suffix.lower()
        mime = {'.jpg':'image/jpeg','.jpeg':'image/jpeg','.png':'image/png',
                '.gif':'image/gif','.webp':'image/webp'}.get(ext,'image/jpeg')

        resp = client.models.generate_content(
            model=MODELO,
            contents=[
                types.Part.from_bytes(data=dados, mime_type=mime),
                PROMPT
            ],
            config=types.GenerateContentConfig(
                tools=[types.Tool(google_search=types.GoogleSearch())],
                http_options=types.HttpOptions(timeout=120000)
            )
        )
        analise = resp.text
        sr      = suc_rev(analise)

        # extrai titulo identificado pelo modelo (primeira linha do texto)
        primeira_linha = analise.strip().split('\n')[0][:100]
        print(f'OK [{sr}] — {primeira_linha}')

        resultados.append({
            'arquivo':          nome_arquivo,
            'analise':          analise,
            'sucesso_revisado': sr,
            'timestamp':        datetime.now().isoformat(),
        })
        time.sleep(4)

    except Exception as e:
        print(f'ERRO: {str(e)[:100]}')
        resultados.append({
            'arquivo':          nome_arquivo,
            'analise':          '',
            'sucesso_revisado': 'erro',
            'timestamp':        datetime.now().isoformat(),
        })

print(f'\nProcessados: {len(resultados)}')

# atualiza JSON — move registros entre principal e sem_ref conforme novo sucesso_revisado
novos_principal = list(principal)
novos_sem_ref   = list(sem_ref)

for res in resultados:
    arq = res['arquivo']

    # remove de onde estiver
    novos_principal = [r for r in novos_principal if r.get('arquivo','') != arq]
    novos_sem_ref   = [r for r in novos_sem_ref   if r.get('arquivo','') != arq]

    # recupera registro original
    reg = arquivo_para_reg_principal.get(arq) or arquivo_para_reg_semref.get(arq)
    if not reg:
        # cria novo registro
        stem = Path(arq).stem
        id_obra = re.sub(r'^[AP]_\s*', '', stem)
        reg = {
            'identificacao':    id_obra,
            'arquivo':          arq,
            'modelo_usado':     MODELO,
            'status':           'sucesso',
            'fonte_dado':       'novo',
        }

    reg['analise']          = res['analise']
    reg['sucesso_revisado'] = res['sucesso_revisado']
    reg['timestamp']        = res['timestamp']

    if res['sucesso_revisado'] == 'sim':
        novos_principal.append(reg)
    else:
        novos_sem_ref.append(reg)

with open(PASTA_OUTPUT / 'beta1.json', 'w', encoding='utf-8') as f:
    json.dump(novos_principal, f, ensure_ascii=False, indent=2)
with open(PASTA_OUTPUT / 'beta1_Obras_Sem_Referencia.json', 'w', encoding='utf-8') as f:
    json.dump(novos_sem_ref, f, ensure_ascii=False, indent=2)
print(f'beta1.json: {len(novos_principal)} | beta1_Obras_Sem_Referencia.json: {len(novos_sem_ref)}')

# atualiza XLSX
wb = openpyxl.load_workbook(PASTA_OUTPUT / 'beta1.xlsx')
ws = wb.active
headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}
col_obra    = headers.get('Obra')
col_arquivo = headers.get('Arquivo')
col_analise = headers.get('Analise')
col_sr      = headers.get('sucesso_revisado')

# mapa arquivo -> linha xlsx (linhas 488-511)
arq_para_linha = {}
for row in range(488, 512):
    val = ws.cell(row=row, column=col_arquivo).value
    if val: arq_para_linha[val] = row

# mapa arquivo -> novo registro
res_map = {r['arquivo']: r for r in resultados}
# mapa arquivo -> identificacao nova (do JSON atualizado)
id_map = {r.get('arquivo',''): r.get('identificacao','') for r in novos_principal + novos_sem_ref}

for arq, linha in arq_para_linha.items():
    res = res_map.get(arq)
    if not res: continue
    novo_id = id_map.get(arq, ws.cell(row=linha, column=col_obra).value)
    ws.cell(row=linha, column=col_obra,    value=novo_id)
    cell = ws.cell(row=linha, column=col_analise, value=res.get('analise',''))
    cell.alignment = Alignment(wrap_text=True)
    ws.cell(row=linha, column=col_sr,      value=res.get('sucesso_revisado',''))

wb.save(PASTA_OUTPUT / 'beta1.xlsx')
print('XLSX salvo.')
