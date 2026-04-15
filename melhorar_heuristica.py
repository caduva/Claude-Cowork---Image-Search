"""
Completa analise_melhorada e sucesso_revisado para obras ainda nao processadas,
usando heuristica (sem LLM). Preserva resultados ja existentes no checkpoint.
"""
import re, sys, json
from pathlib import Path

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

PASTA_OUTPUT    = Path(r"C:\Users\ctucunduva\fontes fuse\Claude Cowork - Image Search\Output Gemini Description")
CHECKPOINT_DIR  = Path(r"C:\Users\ctucunduva\fontes fuse\Claude Cowork - Image Search")
CHECKPOINT_FILE = CHECKPOINT_DIR / "melhorar_descricoes_checkpoint.json"
ARQUIVOS        = ["batch1", "batch2", "batch3", "beta1"]

# frases que indicam que o gemini NAO encontrou referencias
FRASES_NAO_ENCONTROU = [
    'i was unable to find', 'could not find', "couldn't find", 'i could not find',
    'unable to find', 'no specific interpretations', 'no references were found',
    'no information was found', 'no results were found', 'no specific information',
    'did not find', 'no online references', 'not find any', 'unable to locate',
    'no information about', 'no details about', 'no specific details',
]

# frases que indicam referencia explicita a internet (precisam de limpeza)
FRASES_INTERNET = [
    'found on the internet', 'searched the internet', 'search the internet',
    'the search results', 'my search', 'i searched', 'based on my search',
    'according to my search', 'upon searching', 'after searching',
    'from my internet search', 'internet search', 'web search', 'online search',
    'i found', 'search revealed', 'my research revealed', 'search indicates',
    'search shows', 'search provided', 'search returned',
]

# prefixos comuns que introduzem referencia a internet (para remover do inicio)
PREFIXOS_PARA_REMOVER = [
    r"^i was unable to find[^\.]*\.\s*",
    r"^i could not find[^\.]*\.\s*",
    r"^i couldn't find[^\.]*\.\s*",
    r"^unfortunately[^\.]*unable[^\.]*\.\s*",
    r"^the search results (did not|didn't)[^\.]*\.\s*",
    r"^based on (my |the )?search[^\.]*,\s*",
    r"^upon searching[^,\.]*[,\.]\s*",
    r"^after searching[^,\.]*[,\.]\s*",
    r"^my (internet |online |web )?search[^\.]*\.\s*",
]

# substituicoes inline (frase -> versao limpa)
SUBSTITUICOES = [
    (r'\bthe search results (show|indicate|reveal|suggest|provide)[s]?\b', ''),
    (r'\baccording to (my |the )?(internet |online |web )?search[^,]*,\s*', ''),
    (r'\bbased on (my |the )?(internet |online |web )?search[^,]*,\s*', ''),
    (r'\bi found (on the internet|online|through search)[^,]*,?\s*', ''),
    (r'\b(upon|after) (searching|my search)[^,]*,\s*', ''),
    (r'\bmy (internet |online |web )?search revealed\s*', ''),
    (r'\bfound on the internet\b', 'found'),
    (r'\bsearched (the internet|online)\b', ''),
]

def limpar_texto(texto):
    """Remove referencias a internet do texto."""
    t = texto.strip()

    # tenta remover prefixos do inicio
    for padrao in PREFIXOS_PARA_REMOVER:
        novo = re.sub(padrao, '', t, flags=re.IGNORECASE)
        if novo != t:
            t = novo.strip()
            # capitaliza primeira letra
            if t:
                t = t[0].upper() + t[1:]
            break

    # substituicoes inline
    for padrao, subst in SUBSTITUICOES:
        t = re.sub(padrao, subst, t, flags=re.IGNORECASE)

    # limpa espacos duplos
    t = re.sub(r'  +', ' ', t).strip()
    return t

def classificar_sucesso(texto):
    """Classifica se o gemini encontrou informacao real."""
    t = (texto or '').lower().strip()
    if not t:
        return 'nao'

    # verifica se é principalmente "nao encontrou"
    for frase in FRASES_NAO_ENCONTROU:
        if frase in t[:400]:
            # verifica se tem conteudo substantivo alem do "nao encontrou"
            # remove a frase de nao encontrado e ve o que sobra
            restante = t
            for f in FRASES_NAO_ENCONTROU:
                restante = restante.replace(f, '')
            restante = restante.strip()
            if len(restante) > 200:
                return 'parcial'
            return 'nao'

    # texto substantivo sem indicacao de nao encontrado
    if len(t) > 300:
        return 'sim'
    elif len(t) > 100:
        return 'parcial'
    return 'nao'

def tem_referencia_internet(texto):
    t = (texto or '').lower()
    return any(f in t[:800] for f in FRASES_INTERNET)

# carrega checkpoint
with open(CHECKPOINT_FILE, encoding='utf-8') as f:
    checkpoint = json.load(f)

print(f"Checkpoint: {len(checkpoint)} obras ja processadas via LLM\n")

total_heuristica = 0
total_pulado = 0

for nome_base in ARQUIVOS:
    json_path = PASTA_OUTPUT / f"{nome_base}.json"
    xlsx_path = PASTA_OUTPUT / f"{nome_base}.xlsx"

    print(f"\n{'='*60}")
    print(f"Processando: {nome_base}")
    print(f"{'='*60}")

    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)

    for r in data:
        if r.get('status') != 'sucesso':
            r.setdefault('analise_melhorada', '')
            r.setdefault('sucesso_revisado', '')
            continue

        chave = r.get('identificacao', '')
        analise = r.get('analise') or ''

        # ja processado pela LLM — preserva
        if chave in checkpoint:
            r['analise_melhorada'] = checkpoint[chave]['analise_melhorada']
            r['sucesso_revisado']  = checkpoint[chave]['sucesso_revisado']
            total_pulado += 1
            continue

        # heuristica
        if not analise.strip():
            r['analise_melhorada'] = ''
            r['sucesso_revisado']  = 'nao'
        elif tem_referencia_internet(analise):
            r['analise_melhorada'] = limpar_texto(analise)
            r['sucesso_revisado']  = classificar_sucesso(analise)
        else:
            r['analise_melhorada'] = analise  # texto já está limpo
            r['sucesso_revisado']  = classificar_sucesso(analise)

        total_heuristica += 1

    # salva JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON salvo ({len(data)} registros)")

    # atualiza XLSX
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1)}

    if 'analise_melhorada' not in headers:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value='analise_melhorada')
        ws.column_dimensions[get_column_letter(col)].width = 80
        headers['analise_melhorada'] = col

    if 'sucesso_revisado' not in headers:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value='sucesso_revisado')
        ws.column_dimensions[get_column_letter(col)].width = 18
        headers['sucesso_revisado'] = col

    col_id = headers.get('identificacao')
    id_para_linha = {}
    if col_id:
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_id).value
            if val:
                id_para_linha[val] = row

    col_am = headers['analise_melhorada']
    col_sr = headers['sucesso_revisado']
    for r in data:
        chave = r.get('identificacao', '')
        linha = id_para_linha.get(chave)
        if linha:
            cell_am = ws.cell(row=linha, column=col_am, value=r.get('analise_melhorada', ''))
            cell_am.alignment = Alignment(wrap_text=True)
            ws.cell(row=linha, column=col_sr, value=r.get('sucesso_revisado', ''))

    wb.save(xlsx_path)
    print(f"  XLSX salvo")

    # resumo do arquivo
    sim = sum(1 for r in data if r.get('sucesso_revisado') == 'sim')
    parcial = sum(1 for r in data if r.get('sucesso_revisado') == 'parcial')
    nao = sum(1 for r in data if r.get('sucesso_revisado') == 'nao')
    print(f"  sucesso_revisado: sim={sim} | parcial={parcial} | nao={nao}")

print(f"\n{'='*60}")
print(f"CONCLUIDO")
print(f"  Via LLM (preservados): {total_pulado}")
print(f"  Via heuristica:        {total_heuristica}")
print(f"  Total:                 {total_pulado + total_heuristica}")
print(f"{'='*60}")
