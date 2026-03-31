"""
analise_obras_gemini.py
========================
Analisa obras de arte usando Gemini com Google Search Grounding.
O resultado é equivalente ao que você obtém no browser do Gemini:
o modelo busca na internet em tempo real antes de responder.

SETUP (1x — rode no terminal do Windows):
  pip install google-genai openpyxl

COMO USAR:
  1. Defina a variável de ambiente API_KEY_GEMINI_FUSE com sua chave
  2. Execute: python analise_obras_gemini.py

NOME DOS ARQUIVOS:
  O script usa o nome do arquivo como identificação da obra no prompt.
  Exemplo: "Adriana Varejão_Açougue song_2000_(1).jpg"
           → prompt enviado com "Adriana Varejão_Açougue song_2000"

ESTRATÉGIA DE MODELOS:
  1. Tenta primeiro com MODELO_PRIMARIO (gemini-3.1-pro-preview)
  2. Se não encontrar referências, tenta novamente com MODELO_FALLBACK (gemini-3-flash-preview)
"""

import os
import re
import sys
import json
import time
from pathlib import Path
from datetime import datetime

# Força UTF-8 no terminal Windows para evitar UnicodeEncodeError
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ============================================================
#  CONFIG — EDITE AQUI
# ============================================================

API_KEY = os.environ.get("API_KEY_GEMINI_FUSE", "SUA_API_KEY_AQUI")

PASTA_IMAGENS = r"C:\Users\ctucunduva\fontes fuse\images\Additional artworks for Beta 2 (batch 2)"

EXTENSOES = [".jpg", ".jpeg", ".png", ".webp", ".gif", ".JPG", ".PNG", ".JPEG"]

MODELO_PRIMARIO = "gemini-3.1-pro-preview"
MODELO_FALLBACK = "gemini-3-flash-preview"

PROMPT_PADRAO = (
    "Please find interpretations of this artwork on the internet: {identificacao}. "
    "If you can't find the references for the work online, don't try to guess or make "
    "your own interpretation; simply report that you couldn't find them. "
    "Formatting: Do not use bold or italics (no ** or * or _ symbols). "
    "Numbered lists, indentation, and paragraph breaks are allowed."
)

# Frases que indicam que o modelo não encontrou referências
NAO_ENCONTROU_KEYWORDS = [
    "unable to find", "could not find", "couldn't find", "not find",
    "not found", "no references", "no online references", "no interpretations",
    "cannot find", "can't find", "i was unable", "i could not", "i couldn't",
    "did not find", "don't have", "do not have", "no specific information",
    "no details", "no information", "no results", "no specific",
]

PAUSA_ENTRE_REQUISICOES = 3  # Segundos entre chamadas

# ============================================================
#  SCRIPT
# ============================================================

def extrair_identificacao(nome_arquivo):
    """
    Remove prefixo A_ e sufixos numéricos _(1), _1, _text do nome do arquivo.
    Retorna a identificação limpa para o prompt.
    """
    stem = Path(nome_arquivo).stem
    stem = re.sub(r'^A_\s*', '', stem)
    stem = re.sub(r'^A\s+', '', stem)
    stem = re.sub(r'[_\s]+text\s*$', '', stem, flags=re.IGNORECASE)
    stem = re.sub(r'[_\s]+[\(\[]?\d+[\)\]]?\s*$', '', stem)
    return stem.strip()


def agrupar_obras(pasta):
    """
    Agrupa imagens por obra (ignora duplicatas _1, _2, _text).
    Retorna lista de (identificacao, caminho_imagem_representativa).
    """
    arquivos = [
        f for f in Path(pasta).iterdir()
        if f.is_file() and f.suffix in EXTENSOES
    ]
    arquivos.sort()

    grupos = {}
    for f in arquivos:
        if 'text' in f.stem.lower() and f.stem.lower().endswith('text'):
            continue
        chave = extrair_identificacao(f.name)
        if chave not in grupos:
            grupos[chave] = f

    return list(grupos.items())


def configurar_gemini():
    try:
        from google import genai
        client = genai.Client(api_key=API_KEY)
        return client
    except ImportError:
        print("Biblioteca nao instalada. Execute:")
        print("  pip install google-genai openpyxl")
        exit(1)


def nao_encontrou_referencias(texto):
    """Retorna True se o modelo reportou que não encontrou referências online."""
    texto_lower = (texto or "").lower()
    return any(kw in texto_lower for kw in NAO_ENCONTROU_KEYWORDS)


def chamar_modelo(client, caminho_imagem, identificacao, modelo):
    from google.genai import types

    with open(caminho_imagem, "rb") as f:
        dados_imagem = f.read()

    ext = Path(caminho_imagem).suffix.lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".png": "image/png", ".webp": "image/webp",
                ".gif": "image/gif"}
    mime_type = mime_map.get(ext, "image/jpeg")

    prompt = PROMPT_PADRAO.format(identificacao=identificacao)

    response = client.models.generate_content(
        model=modelo,
        contents=[
            types.Part.from_bytes(data=dados_imagem, mime_type=mime_type),
            prompt,
        ],
        config=types.GenerateContentConfig(
            tools=[types.Tool(google_search=types.GoogleSearch())]
        )
    )

    return response.text


def analisar_obra(client, caminho_imagem, identificacao):
    """
    Tenta primeiro com o modelo primário (3.1 Pro).
    Se não encontrar referências, tenta com o fallback (3 Flash).
    Retorna (texto, modelo_usado).
    """
    texto = chamar_modelo(client, caminho_imagem, identificacao, MODELO_PRIMARIO)

    if nao_encontrou_referencias(texto):
        time.sleep(PAUSA_ENTRE_REQUISICOES)
        texto_fallback = chamar_modelo(client, caminho_imagem, identificacao, MODELO_FALLBACK)
        # Sempre usa o Flash quando o Pro não encontrou:
        # - Se Flash encontrou → melhor resultado
        # - Se Flash também não encontrou → resposta honesta, sem alucinação híbrida do Pro
        return texto_fallback, MODELO_FALLBACK

    return texto, MODELO_PRIMARIO


def salvar_json(resultados, caminho_saida):
    with open(caminho_saida, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    print(f"JSON salvo: {caminho_saida}")


def salvar_excel(resultados, caminho_saida):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl nao instalado. Execute: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analise de Obras"

    cabecalho = ["#", "Identificacao da Obra", "Arquivo", "Analise Gemini", "Modelo Usado", "Status", "Timestamp"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    fallback_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for i, item in enumerate(resultados, 1):
        row = i + 1
        usou_fallback = item.get("modelo_usado") == MODELO_FALLBACK
        fill = fallback_fill if usou_fallback else (alt_fill if i % 2 == 0 else PatternFill())

        valores = [
            i,
            item.get("identificacao", ""),
            item.get("arquivo", ""),
            item.get("analise", item.get("erro", "")),
            item.get("modelo_usado", ""),
            item.get("status", ""),
            item.get("timestamp", ""),
        ]

        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    larguras = [5, 45, 35, 100, 22, 10, 20]
    for col, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    for row in range(2, len(resultados) + 2):
        ws.row_dimensions[row].height = 120

    ws.freeze_panes = "A2"
    wb.save(caminho_saida)
    print(f"Excel salvo: {caminho_saida}")


def main():
    print("=" * 60)
    print("  ANALISE DE OBRAS — GEMINI + GOOGLE SEARCH GROUNDING")
    print("=" * 60)

    if API_KEY == "SUA_API_KEY_AQUI":
        print("\nERRO: Configure sua API key no inicio do script.")
        exit(1)

    pasta = Path(PASTA_IMAGENS)
    if not pasta.exists():
        print(f"\nERRO: Pasta nao encontrada: {PASTA_IMAGENS}")
        exit(1)

    obras = agrupar_obras(PASTA_IMAGENS)
    if not obras:
        print(f"\nERRO: Nenhuma imagem encontrada em: {PASTA_IMAGENS}")
        exit(1)

    print(f"\n{len(obras)} obras unicas encontradas")
    print(f"Modelo primario:  {MODELO_PRIMARIO}")
    print(f"Modelo fallback:  {MODELO_FALLBACK}\n")

    client = configurar_gemini()
    pasta_script = Path(__file__).parent

    # Retoma do checkpoint mais recente, se existir
    resultados = []
    ja_processados = set()
    checkpoints = sorted(pasta_script.glob("analise_obras_*_parcial.json"), reverse=True)
    timestamp_inicio = datetime.now().strftime("%Y%m%d_%H%M%S")
    if checkpoints:
        cp_mais_recente = checkpoints[0]
        with open(cp_mais_recente, encoding="utf-8") as f:
            resultados = json.load(f)
        # Só pula obras com sucesso — reprocessa erros
        ja_processados = {r["identificacao"] for r in resultados if r.get("status") == "sucesso"}
        # Remove erros do resultados para que sejam reprocessados e substituídos
        resultados = [r for r in resultados if r.get("status") == "sucesso"]
        timestamp_inicio = cp_mais_recente.stem.replace("analise_obras_", "").replace("_parcial", "")
        print(f"Retomando do checkpoint: {cp_mais_recente.name} ({len(resultados)} com sucesso, erros serao reprocessados)\n")

    for idx, (identificacao, caminho) in enumerate(obras, 1):
        if identificacao in ja_processados:
            continue
        print(f"[{idx}/{len(obras)}] {identificacao[:70]}...", end=" ", flush=True)

        try:
            analise, modelo_usado = analisar_obra(client, str(caminho), identificacao)
            flag = f"[fallback: {MODELO_FALLBACK}]" if modelo_usado == MODELO_FALLBACK else ""
            resultados.append({
                "id": idx,
                "identificacao": identificacao,
                "arquivo": caminho.name,
                "analise": analise,
                "modelo_usado": modelo_usado,
                "status": "sucesso",
                "timestamp": datetime.now().isoformat(),
            })
            print(f"OK {flag}")

        except Exception as e:
            resultados.append({
                "id": idx,
                "identificacao": identificacao,
                "arquivo": caminho.name,
                "analise": "",
                "erro": str(e),
                "modelo_usado": "",
                "status": "erro",
                "timestamp": datetime.now().isoformat(),
            })
            print(f"ERRO: {e}")

        if idx % 5 == 0:
            cp = pasta_script / f"analise_obras_{timestamp_inicio}_parcial.json"
            with open(cp, "w", encoding="utf-8") as f:
                json.dump(resultados, f, ensure_ascii=False, indent=2)
            print(f"  >> Checkpoint salvo ({idx} obras)")

        if idx < len(obras):
            time.sleep(PAUSA_ENTRE_REQUISICOES)

    print("\n" + "=" * 60)
    saida_json = pasta_script / f"analise_obras_{timestamp_inicio}.json"
    saida_xlsx = pasta_script / f"analise_obras_{timestamp_inicio}.xlsx"

    salvar_json(resultados, str(saida_json))
    salvar_excel(resultados, str(saida_xlsx))

    sucessos = sum(1 for r in resultados if r["status"] == "sucesso")
    fallbacks = sum(1 for r in resultados if r.get("modelo_usado") == MODELO_FALLBACK)
    print(f"\nRESUMO: {sucessos}/{len(resultados)} obras processadas com sucesso")
    print(f"         {fallbacks} obras resolvidas pelo fallback ({MODELO_FALLBACK})")
    print(f"Arquivos salvos em: {pasta_script}")


if __name__ == "__main__":
    main()
